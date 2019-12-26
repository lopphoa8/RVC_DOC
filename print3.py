# -*- coding: utf-8 -*-
import openpyxl
import time

file1 = 'MarketingAnalystNames.xlsm'
file2 = 'MarketingAnalystNames_Merge.xlsm'
file3 = 'MarketingAnalystNames_Merge.xlsm'

start_time = time.time()

#wb1 = openpyxl.load_workbook(filename=file1, read_only=False, keep_vba=True)
#wb2 = openpyxl.load_workbook(filename=file2, read_only=False, keep_vba=True)

wb1 = openpyxl.load_workbook(filename=file1, read_only=True, keep_vba=False)
wb2 = openpyxl.load_workbook(filename=file2, keep_vba=True)

print("--- %s seconds ---" % (time.time() - start_time))

#print(len(wb1.sheetnames))
sheets = wb1.sheetnames

for sh in range (len(wb1.sheetnames)):
        sheet1 = wb1.get_sheet_by_name(sheets[sh])
        sheet2 = wb2.get_sheet_by_name(sheets[sh])
        for i in range(1,sheet1.max_row+1):
                for j in range(1,sheet1.max_column+1):
                                #print(sh,i,j,sheet1.cell(i,j).value)
                                sheet2.cell(i,j).value = sheet1.cell(i,j).value
                                #print(sh,i,j,sheet2.cell(i,j).value)

                                

# Escribirmos en el Fichero

print("--- %s seconds ---" % (time.time() - start_time))

#wb1.save(file1)
#wb2.save(file2)
wb2.save(file3)

print("--- %s seconds ---" % (time.time() - start_time))
